# -*- coding: utf-8 -*-
"""Reordena la composicion de saldos de cta cte de proveedores.
Lee el .xls original (SIN modificarlo) y genera un .xlsx limpio con:
  - Hoja 'Detalle Ordenado': columnas separadas + Debe/Haber + saldo por proveedor coloreado.
  - Hoja 'Resumen Proveedores': un renglon por proveedor con estado coloreado.
  - Hoja 'Estado por Factura': concilia cada comprobante con sus pagos.
"""
import re
import shutil
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\tomas\Desktop\adcg - scg - composicion de saldos de ctacte detallado por referencia.xls"
COPIA_XLS = r"C:\Users\tomas\Desktop\COPIA - composicion de saldos (original sin tocar).xls"
OUT = r"C:\Users\tomas\Desktop\composicion de saldos - ORDENADO v2.xlsx"

# 1) Copia de resguardo del original (no se modifica el original)
shutil.copyfile(SRC, COPIA_XLS)

wb = xlrd.open_workbook(SRC)
sh = wb.sheet_by_index(0)
datemode = wb.datemode


def to_date(val):
    """Convierte serial de Excel a datetime; devuelve None si no aplica."""
    if isinstance(val, float) and val > 20000:
        try:
            return xlrd.xldate_as_datetime(val, datemode)
        except Exception:
            return None
    return None


def parse_comentario(c):
    """Devuelve (tipo, comprobante, referencia_factura, base).
    - FC/NC: 'FC-C-00002-00000223/[Sin Comentario]'
    - OP:    'OP-000377 / FC-C-00002-00000223'
    base = documento de factura al que pertenece el renglon (para conciliar)."""
    c = (c or "").strip()
    if not c:
        return ("", "", "", "")
    tipo = c[:2].upper()
    if tipo == "OP":
        partes = c.split("/", 1)
        nro_op = partes[0].replace("OP-", "").strip()
        ref = partes[1].strip() if len(partes) > 1 else ""
        return ("OP", nro_op, ref, ref)  # base = factura que paga
    if tipo in ("FC", "NC"):
        izq = c.split("/", 1)[0].strip()          # FC-C-00002-00000223
        comprobante = izq[3:] if len(izq) > 3 else izq  # C-00002-00000223
        return (tipo, comprobante, "", izq)        # base = su propio comprobante
    return ("", c, "", c)


# 2) Recorremos filas agrupando por proveedor (PO)
proveedores = []          # lista de dicts {po, nombre, filas:[...]}
actual = None
for r in range(sh.nrows):
    a = sh.cell_value(r, 0)
    b = sh.cell_value(r, 1)
    com = sh.cell_value(r, 2)
    ncta = sh.cell_value(r, 3)
    nomcta = sh.cell_value(r, 4)
    imp = sh.cell_value(r, 5)

    a_str = str(a).strip()

    # fila de encabezado de grupo: "PO xxxxx" + nombre en col B
    if a_str.upper().startswith("PO ") or re.match(r"^PO\s*\d", a_str, re.I):
        actual = {"po": a_str.strip(), "nombre": str(b).strip(), "filas": []}
        proveedores.append(actual)
        continue

    # fila de transaccion valida: tiene fecha serial y un importe numerico
    fecha = to_date(a)
    if fecha is None:
        continue
    if not isinstance(imp, (int, float)) or imp == "":
        continue
    if actual is None:
        continue

    tipo, comprobante, ref, base = parse_comentario(com)
    asiento = b
    if isinstance(asiento, float):
        asiento = int(asiento) if asiento == int(asiento) else asiento
    ncta_v = ncta
    if isinstance(ncta_v, float):
        ncta_v = int(ncta_v) if ncta_v == int(ncta_v) else ncta_v

    debe = round(-imp, 2) if imp < 0 else 0.0      # facturas (deuda) -> positivo
    haber = round(imp, 2) if imp > 0 else 0.0       # pagos / NC        -> positivo

    actual["filas"].append({
        "fecha": fecha,
        "asiento": asiento,
        "tipo": tipo,
        "comprobante": comprobante,
        "referencia": ref,
        "base": base,
        "ncta": ncta_v,
        "cuenta": str(nomcta).strip(),
        "importe": round(imp, 2),
        "debe": debe,
        "haber": haber,
    })

# quitar proveedores sin transacciones
proveedores = [p for p in proveedores if p["filas"]]
print("Proveedores con movimientos:", len(proveedores))
print("Total renglones:", sum(len(p["filas"]) for p in proveedores))

# ----------------------------------------------------------------------------
# 3) Construir el .xlsx
# ----------------------------------------------------------------------------
wbo = Workbook()

# --- estilos ---
FONT_TIT = Font(bold=True, size=14, color="FFFFFF")
FONT_HDR = Font(bold=True, color="FFFFFF")
FONT_GRP = Font(bold=True, color="1F3864")
FONT_SALDO = Font(bold=True)
FILL_TIT = PatternFill("solid", fgColor="1F3864")
FILL_HDR = PatternFill("solid", fgColor="2F5496")
FILL_GRP = PatternFill("solid", fgColor="D6E4F0")
FILL_VERDE = PatternFill("solid", fgColor="C6EFCE")   # pagado
FILL_ROJO = PatternFill("solid", fgColor="FFC7CE")    # a pagar
FILL_AMAR = PatternFill("solid", fgColor="FFEB9C")    # a favor / saldo positivo
FILL_ZEBRA = PatternFill("solid", fgColor="F2F6FB")
BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")
MONEY = '#,##0.00'
DATEF = 'dd/mm/yyyy'

EPS = 0.5  # tolerancia para considerar saldo = 0


def estado_saldo(saldo):
    if abs(saldo) <= EPS:
        return ("PAGADO", FILL_VERDE)
    if saldo < 0:
        return ("A PAGAR", FILL_ROJO)
    return ("SALDO A FAVOR", FILL_AMAR)


# =========================== HOJA 1: DETALLE ===========================
ws = wbo.active
ws.title = "Detalle Ordenado"
cols = ["Codigo PO", "Proveedor", "Fecha", "Asiento", "Tipo",
        "Nro Comprobante", "Factura Referenciada", "Nro Cuenta", "Cuenta",
        "Debe (Facturas)", "Haber (Pagos/NC)", "Importe Neto", "Estado"]
widths = [12, 34, 12, 10, 7, 20, 24, 12, 14, 16, 16, 15, 15]

# titulo
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
tc = ws.cell(1, 1, "COMPOSICION DE SALDOS DE CTA CTE - PROVEEDORES (ordenado)")
tc.font = FONT_TIT
tc.fill = FILL_TIT
tc.alignment = CENTER
ws.row_dimensions[1].height = 24

# encabezado
hr = 2
for j, name in enumerate(cols, 1):
    c = ws.cell(hr, j, name)
    c.font = FONT_HDR
    c.fill = FILL_HDR
    c.alignment = CENTER
    c.border = BORDER
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A3"

row = hr + 1
for p in proveedores:
    saldo = round(sum(f["importe"] for f in p["filas"]), 2)
    tot_debe = round(sum(f["debe"] for f in p["filas"]), 2)
    tot_haber = round(sum(f["haber"] for f in p["filas"]), 2)
    estado, fill_estado = estado_saldo(saldo)

    # fila de grupo del proveedor
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    gc = ws.cell(row, 1, f"{p['po']}   -   {p['nombre']}")
    gc.font = FONT_GRP
    gc.fill = FILL_GRP
    for j in range(1, len(cols) + 1):
        ws.cell(row, j).fill = FILL_GRP
    row += 1

    for i, f in enumerate(p["filas"]):
        vals = [p["po"], p["nombre"], f["fecha"], f["asiento"], f["tipo"],
                f["comprobante"], f["referencia"], f["ncta"], f["cuenta"],
                f["debe"] or None, f["haber"] or None, f["importe"], None]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.border = BORDER
            if i % 2 == 1:
                c.fill = FILL_ZEBRA
        ws.cell(row, 3).number_format = DATEF
        for j in (10, 11, 12):
            ws.cell(row, j).number_format = MONEY
        row += 1

    # subtotal / saldo del proveedor
    sc = ws.cell(row, 9, "SALDO PROVEEDOR:")
    sc.font = FONT_SALDO
    sc.alignment = RIGHT
    ws.cell(row, 10, tot_debe).number_format = MONEY
    ws.cell(row, 11, tot_haber).number_format = MONEY
    ws.cell(row, 12, saldo).number_format = MONEY
    ec = ws.cell(row, 13, estado)
    ec.alignment = CENTER
    for j in (10, 11, 12, 13):
        ws.cell(row, j).font = FONT_SALDO
        ws.cell(row, j).fill = fill_estado
        ws.cell(row, j).border = BORDER
    row += 2  # espacio entre proveedores

# =========================== HOJA 2: RESUMEN ===========================
ws2 = wbo.create_sheet("Resumen Proveedores")
cols2 = ["Codigo PO", "Proveedor", "Total Facturado", "Total Pagado/NC",
         "Saldo", "Estado", "Cant. Movim."]
w2 = [12, 40, 18, 18, 16, 16, 13]
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols2))
t2 = ws2.cell(1, 1, "RESUMEN POR PROVEEDOR")
t2.font = FONT_TIT; t2.fill = FILL_TIT; t2.alignment = CENTER
ws2.row_dimensions[1].height = 22
for j, name in enumerate(cols2, 1):
    c = ws2.cell(2, j, name); c.font = FONT_HDR; c.fill = FILL_HDR
    c.alignment = CENTER; c.border = BORDER
for j, w in enumerate(w2, 1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.freeze_panes = "A3"

r2 = 3
tot_deb_g = tot_hab_g = tot_sal_g = 0.0
# ordenar: primero los que hay que pagar (saldo mas negativo primero)
prov_sorted = sorted(proveedores, key=lambda p: sum(f["importe"] for f in p["filas"]))
for p in prov_sorted:
    saldo = round(sum(f["importe"] for f in p["filas"]), 2)
    td = round(sum(f["debe"] for f in p["filas"]), 2)
    thh = round(sum(f["haber"] for f in p["filas"]), 2)
    estado, fill_estado = estado_saldo(saldo)
    tot_deb_g += td; tot_hab_g += thh; tot_sal_g += saldo
    vals = [p["po"], p["nombre"], td, thh, saldo, estado, len(p["filas"])]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(r2, j, v); c.border = BORDER
    for j in (3, 4, 5):
        ws2.cell(r2, j).number_format = MONEY
    for j in (5, 6):
        ws2.cell(r2, j).fill = fill_estado
        ws2.cell(r2, j).font = FONT_SALDO
    ws2.cell(r2, 6).alignment = CENTER
    ws2.cell(r2, 7).alignment = CENTER
    r2 += 1
# totales generales
tc = ws2.cell(r2, 2, "TOTALES GENERALES"); tc.font = FONT_SALDO; tc.alignment = RIGHT
ws2.cell(r2, 3, round(tot_deb_g, 2)).number_format = MONEY
ws2.cell(r2, 4, round(tot_hab_g, 2)).number_format = MONEY
ws2.cell(r2, 5, round(tot_sal_g, 2)).number_format = MONEY
for j in (3, 4, 5):
    ws2.cell(r2, j).font = FONT_SALDO
    ws2.cell(r2, j).fill = FILL_GRP

# =========================== HOJA 3: ESTADO POR FACTURA ===========================
ws3 = wbo.create_sheet("Estado por Factura")
cols3 = ["Codigo PO", "Proveedor", "Fecha", "Mes", "Documento (Factura/NC)",
         "Facturado", "Pagado aplicado", "Saldo", "Estado"]
w3 = [12, 38, 12, 12, 26, 16, 16, 15, 14]
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols3))
t3 = ws3.cell(1, 1, "ESTADO POR COMPROBANTE (concilia factura vs pagos)")
t3.font = FONT_TIT; t3.fill = FILL_TIT; t3.alignment = CENTER
ws3.row_dimensions[1].height = 22
for j, name in enumerate(cols3, 1):
    c = ws3.cell(2, j, name); c.font = FONT_HDR; c.fill = FILL_HDR
    c.alignment = CENTER; c.border = BORDER
for j, w in enumerate(w3, 1):
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.freeze_panes = "A3"

r3 = 3
for p in proveedores:
    # agrupar por documento base
    docs = {}
    orden = []
    for f in p["filas"]:
        base = f["base"] or "(sin ref)"
        if base not in docs:
            docs[base] = {"facturado": 0.0, "pagado": 0.0,
                          "fecha_factura": None, "fecha_min": None}
            orden.append(base)
        d = docs[base]
        if f["tipo"] in ("FC", "NC"):
            d["facturado"] += f["importe"]
            # fecha del comprobante = fecha de la factura/NC original
            if d["fecha_factura"] is None or f["fecha"] < d["fecha_factura"]:
                d["fecha_factura"] = f["fecha"]
        else:  # OP u otros pagos
            d["pagado"] += f["importe"]
        # fallback: primera fecha de cualquier movimiento del documento
        if d["fecha_min"] is None or f["fecha"] < d["fecha_min"]:
            d["fecha_min"] = f["fecha"]
    MESES = ["", "01-Enero", "02-Febrero", "03-Marzo", "04-Abril", "05-Mayo",
             "06-Junio", "07-Julio", "08-Agosto", "09-Septiembre",
             "10-Octubre", "11-Noviembre", "12-Diciembre"]
    for base in orden:
        d = docs[base]
        fact = round(d["facturado"], 2)
        pag = round(d["pagado"], 2)
        saldo = round(fact + pag, 2)
        estado, fill_estado = estado_saldo(saldo)
        fecha = d["fecha_factura"] or d["fecha_min"]
        mes = MESES[fecha.month] if fecha else ""
        vals = [p["po"], p["nombre"], fecha, mes, base, fact, pag, saldo, estado]
        for j, v in enumerate(vals, 1):
            c = ws3.cell(r3, j, v); c.border = BORDER
        ws3.cell(r3, 3).number_format = DATEF
        for j in (6, 7, 8):
            ws3.cell(r3, j).number_format = MONEY
        for j in (8, 9):
            ws3.cell(r3, j).fill = fill_estado
            ws3.cell(r3, j).font = FONT_SALDO
        ws3.cell(r3, 9).alignment = CENTER
        r3 += 1

# autofiltros para filtrar por mes / estado / proveedor
ws3.auto_filter.ref = f"A2:I{r3 - 1}"
ws2.auto_filter.ref = f"A2:G{r2}"

wbo.save(OUT)
print("Copia de resguardo:", COPIA_XLS)
print("Archivo ordenado:  ", OUT)
